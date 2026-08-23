# -*- coding: utf-8 -*-
"""B1 부속 엑셀 2종
  ① 가정값_레지스트리.xlsx      — D4 부속 · 22개 가정값 + 폐기 6종
  ② Golden_페르소나검토.xlsx     — D5 부속 · 수동 검토 28건 시트
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

F = 'Arial'
HDR  = PatternFill('solid', fgColor='1F3864')
IN   = PatternFill('solid', fgColor='FFFF00')
AUTO = PatternFill('solid', fgColor='E8E8E8')
SEC  = PatternFill('solid', fgColor='D9E2F3')
WARN = PatternFill('solid', fgColor='FFE0E0')
OK   = PatternFill('solid', fgColor='E2EFDA')
THIN = Side(style='thin', color='B0B0B0')
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def hdr_row(ws, r, cols):
    for i, (t, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        c = ws.cell(r, i, t)
        c.font = Font(F, 9, bold=True, color='FFFFFF')
        c.fill = HDR; c.border = BOX
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[r].height = 30


def put(ws, r, vals, fills=None, wrap=None, sz=9):
    for i, v in enumerate(vals, 1):
        c = ws.cell(r, i, v)
        c.font = Font(F, sz); c.border = BOX
        c.alignment = Alignment(vertical='center', wrap_text=bool(wrap and i in wrap))
        if fills and i in fills: c.fill = fills[i]
    ws.row_dimensions[r].height = 26


# ══════════════════ ① 가정값 레지스트리 ══════════════════
def build_assumptions(path):
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # ── 시트 1: 가정값 ──
    ws = wb.create_sheet('가정값')
    ws.sheet_view.showGridLines = False
    ws.cell(1, 1, 'CREDEAL 가정값 레지스트리 v1.0').font = Font(F, 14, bold=True, color='1F3864')
    ws.cell(2, 1, '노란색 셀만 편집합니다. reviewedAt은 연 1회 갱신 대상입니다.').font = Font(F, 9, italic=True)

    cols = [('key', 26), ('값', 15), ('단위', 9), ('source', 14), ('신뢰도', 8),
            ('편집가능', 9), ('basis (화면 노출)', 46), ('틀리면 무엇이 어긋나나', 44), ('reviewedAt', 12)]
    hdr_row(ws, 4, cols)

    rows = [
        # legal 5
        ('acquisitionTaxRate', 0.046, '비율', 'legal', 'high', 'N',
         '취득세 4.0 + 지방교육세 0.4 + 농특세 0.2 (상가·업무시설 표준세율)',
         '총취득원가·실투자금이 매매가의 4.6%만큼 어긋남', '2026-08-23'),
        ('brokerFeeRateMax', 0.009, '비율', 'legal', 'high', 'N',
         '법정 상한 · 실제는 협의',
         '총취득원가 과대 추정', '2026-08-23'),
        ('targetFarByZoning', None, '%', 'legal', 'high', 'Y',
         '용도지역별 법정 상한 · 토지이용계획 API 조회 · 실패 시 산출 거부',
         '★ 신축 규모·분양수입·사업이익률 전부 어긋남 (400% 고정 시 +60% 과대)', '2026-08-23'),
        ('bcrByZoning', None, '%', 'legal', 'high', 'Y',
         '용도지역별 건폐율 상한 · 동일 API',
         '건축 가능 면적 오판', '2026-08-23'),
        ('transferTaxRate', None, '비율', 'legal', 'high', 'Y',
         '1년 미만 50% / 그 외 6~45% · 주택 단기중과 미적용 · 매수 주체별 상이',
         '매매형 세후 차익이 최대 2배 차이', '2026-08-23'),
        # market_default 8
        ('constructionCostPerPyeong', 12000000, '원/평', 'market_default', 'medium', 'Y',
         '서울 소형 근생 신축 2026 통상 단가 (잠원동 실매물 IM 실측)',
         '★ 총사업비 직결 · 614평에서 1만원 오차 = 614만원', '2026-08-23'),
        ('devContingencyRate', 0.05, '비율', 'market_default', 'medium', 'Y',
         '총사업비 대비 예비비 · 통상 3~7%',
         '사업이익률 소폭 변동', '2026-08-23'),
        ('loanRateDefault', 0.045, '연', 'market_default', 'medium', 'Y',
         '2026 상업용 담보대출 통상 금리',
         '★ 역레버리지 판정이 뒤집힘 (수익률 2.24% 물건은 경계에 민감)', '2026-08-23'),
        ('ltvScenarios', '0 / 40 / 50', '%', 'market_default', 'high', 'Y',
         '표준 제시 3안',
         '시나리오 범위가 실제 대출 조건과 괴리', '2026-08-23'),
        ('pfEquityRatio2026', 0.10, '비율', 'market_default', 'high', 'N',
         'PF 자기자본비율 규제 · 2027년 15% · 2028년 20%',
         '개발형 필요 자기자본 오산', '2026-08-23'),
        ('depreciationYears', 40, '년', 'market_default', 'low', 'Y',
         '철근콘크리트 정액 · 세무 확인 필요',
         '사옥형 절세 효과 추정치 변동', '2026-08-23'),
        ('buildingValueRatio', 0.35, '비율', 'market_default', 'low', 'Y',
         '매매가 중 건물분 비중 · 실제 20~50% 편차',
         '★ 사옥형 세후 판단이 반전될 수 있음', '2026-08-23'),
        ('seoulHotelRevPar', 207345, '원', 'market_default', 'medium', 'Y',
         '서울 호텔 2025 평균 · 4~5성급 포함 · 3성급은 상당히 낮음',
         '운영형 GOP 역산이 등급별로 크게 다름', '2026-08-23'),
        # user_input 6
        ('opexKrw', None, '원/년', 'user_input', 'high', 'Y',
         '실제 운영비 · 미입력 시 NOI 계열 산출 안 함',
         'NOI 4종 미산출 (gross 계열만 제공)', '2026-08-23'),
        ('gopMarginPct', None, '%', 'user_input', 'high', 'Y',
         '운영 실적 기반 · 미입력 시 GOP 산출 안 함',
         'GOP · GOP Cap Rate 미산출', '2026-08-23'),
        ('manualComps', None, '건', 'user_input', 'high', 'Y',
         '중개인 조사 비교사례 · 20억 미만·300억 초과 물건은 필수',
         '★ 목표 매각가·시세갭 미산출 (창작 금지)', '2026-08-23'),
        ('marketRentPerPyeong', None, '원/평', 'user_input', 'high', 'Y',
         '해당 지역 사무실 임차 시세 · 사옥형 절감액 산출용',
         '사옥형 절감 임차료 미산출 → 실질 부담 판단 불가', '2026-08-23'),
        ('appraisedValueKrw', None, '원', 'user_input', 'medium', 'Y',
         '은행 감정가 · 실제 대출 가능액 기준',
         '대출 가능액이 매매가 기준으로 과대 추정', '2026-08-23'),
        ('firstContractDate', None, '날짜', 'user_input', 'high', 'Y',
         '임차인 최초 입주일 · 상가 갱신요구권 10년 기산',
         '★ 갱신요구권 잔여 미산출 → 명도 시점 판단 불가', '2026-08-23'),
        # 규제 시한 2
        ('regulationBasis', '서울시 소규모 건축물 용적률 완화', '—', 'legal', 'high', 'Y',
         '제2종 200→250% · 제3종 250→300%',
         '완화 미적용 시 신축 연면적 19% 감소', '2026-08-23'),
        ('regulationExpiry', '2028-05-18', '날짜', 'legal', 'high', 'N',
         '3년 한시 (2025-05-19 시행)',
         '★ 기한 미표기 시 매수인이 인허가 시점을 오판', '2026-08-23'),
    ]
    r = 5
    for k, v, u, src, conf, ed, basis, impact, rev in rows:
        fills = {2: IN, 7: IN, 9: IN}
        if src == 'legal' and v is None: fills[2] = WARN
        if src == 'user_input': fills[2] = WARN
        if conf == 'low': fills[5] = WARN
        put(ws, r, [k, v if v is not None else 'null (기본값 없음)', u, src, conf, ed, basis, impact, rev],
            fills=fills, wrap={7, 8})
        r += 1
    ws.freeze_panes = ws.cell(5, 1)

    dv = DataValidation(type='list', formula1='"Y,N"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv); dv.add(f'F5:F{r-1}')

    print(f'  가정값 {len(rows)}종')

    # ── 시트 2: 폐기 상수 ──
    ws2 = wb.create_sheet('폐기상수')
    ws2.sheet_view.showGridLines = False
    ws2.cell(1, 1, '폐기 대상 상수 6종 — 단계 2 완료 시 코드에서 제거').font = Font(F, 14, bold=True, color='C00000')
    cols2 = [('폐기 상수', 26), ('현행 값', 16), ('폐기 사유', 46), ('대체', 30), ('제거 확인', 11)]
    hdr_row(ws2, 3, cols2)
    dep = [
        ('NOI 추정 계수', '0.85', '근거 없음 · 문서·주석 전무 · 임의 상수 확정', 'gross 계열만 산출', ''),
        ('Opex Ratio 6종', '12~35%', '출처 없음 · 호텔 35%는 GOP 마진과 혼동 의심', 'opexKrw 실입력', ''),
        ('개발형 용적률', '400%', '용도지역 무시 · 잠원동 적용 시 +60% 과대', 'targetFarByZoning', ''),
        ('개발형 공사비', '800만원/평', '실물 1,200만원 · 33% 과소 · 614평에서 24.56억 차이', 'constructionCostPerPyeong', ''),
        ('Trading 목표 매각가', '매입가 × 1.2', 'comps 없이 차익 23억 창작 (당산동 기준)', 'manualComps 필수', ''),
        ('Trading 비교사례', '매입평당가 × 1.15', '동일 — 근거 없는 할인율 자동 생성', 'manualComps 필수', ''),
    ]
    r = 4
    for a, b, c, d, e in dep:
        put(ws2, r, [a, b, c, d, e], fills={2: WARN, 5: IN}, wrap={3})
        r += 1
    ws2.cell(r + 1, 1, '검증 명령').font = Font(F, 10, bold=True)
    ws2.cell(r + 2, 1, r'rg -n "0\.85|opexRatio|400.*용적률|\* 1\.2\b|\* 1\.15\b" src/domain/building/mobile-im/').font = Font('Consolas', 9)
    ws2.cell(r + 3, 1, '→ 하나라도 걸리면 단계 2 DoD 미충족').font = Font(F, 9, color='C00000')

    # ── 시트 3: 갱신 이력 ──
    ws3 = wb.create_sheet('갱신이력')
    ws3.sheet_view.showGridLines = False
    ws3.cell(1, 1, '가정값 변경 이력').font = Font(F, 14, bold=True, color='1F3864')
    ws3.cell(2, 1, 'market_default 8종은 연 1회(매년 1월) 갱신 대상입니다.').font = Font(F, 9, italic=True)
    hdr_row(ws3, 4, [('일자', 12), ('key', 26), ('이전 값', 14), ('새 값', 14),
                     ('사유', 40), ('승인자', 14)])
    put(ws3, 5, ['2026-08-23', '(초기 등록)', '—', '—', 'D4 레지스트리 최초 작성', ''],
        fills={6: IN}, wrap={5})
    for i in range(6, 26):
        put(ws3, i, ['', '', '', '', '', ''], fills={c: IN for c in range(1, 7)})

    wb.save(path)
    print(f'saved {path}')


# ══════════════════ ② Golden 페르소나 검토 ══════════════════
def build_golden_review(path):
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet('검토절차')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 78
    ws.cell(2, 2, 'Golden Set 정제 — 수동 검토 28건').font = Font(F, 14, bold=True, color='1F3864')

    def sec(r, t):
        for c in (2, 3):
            x = ws.cell(r, c, t if c == 2 else '')
            x.font = Font(F, 11, bold=True, color='FFFFFF'); x.fill = HDR
    def kv(r, k, v, warn=False):
        a = ws.cell(r, 2, k); a.font = Font(F, 10, bold=True); a.border = BOX
        b = ws.cell(r, 3, v); b.font = Font(F, 10); b.border = BOX
        b.alignment = Alignment(wrap_text=True, vertical='center')
        if warn: a.fill = WARN; b.fill = WARN
        ws.row_dimensions[r].height = 32

    r = 4; sec(r, '오염 실측 (164건 중 154건 = 93.9%)'); r += 1
    for k, v in [('이모지 잔여', '128건 (78.0%) — 자동 정제'),
                 ('페르소나 누수', '28건 (17.1%) — 이 시트에서 수동 검토'),
                 ('중복 markdown', '13건 (7.9%) — 자동 정제'),
                 ('가짜 데이터', '0건 — 사실 오류는 축적되지 않았습니다'),
                 ('금지어', '0건')]:
        kv(r, k, v); r += 1

    r += 1; sec(r, '페르소나 판정 기준'); r += 1
    for k, v in [('제거 대상', '"60대 자산가를 위한" · "법인 대표 맞춤" · "초보 투자자용" 등 '
                             '연령·계층·경험 수준을 직접 지칭하는 표현'),
                 ('유지 대상', '"임대수익형 투자자" · "사옥 수요 법인" 등 '
                             '투자 목적·주체 유형 서술은 페르소나 지칭이 아닙니다'),
                 ('판정 애매', '보류로 두고 도메인 담당 확인')]:
        kv(r, k, v); r += 1

    r += 1; sec(r, '주의'); r += 1
    kv(r, '자동 정제 먼저', '이모지·중복 141건을 스크립트로 처리한 뒤 이 시트를 작성합니다.', warn=True); r += 1
    kv(r, '원본 백업', '정제 전 markdown을 별도 컬럼에 보관합니다. 롤백 경로입니다.', warn=True); r += 1
    kv(r, '근본 대책', 'sanitizePersona·stripMarkdown을 Golden 저장 경로에 삽입해야 '
                    '재발하지 않습니다 (응급 E4).', warn=True); r += 1

    # 검토 시트
    ws2 = wb.create_sheet('검토목록')
    ws2.sheet_view.showGridLines = False
    ws2.cell(1, 1, 'Golden 페르소나 수동 검토 — 28건').font = Font(F, 14, bold=True, color='1F3864')
    ws2.cell(2, 1, '노란색 셀만 입력합니다. 판정 후 is_active를 반영하십시오.').font = Font(F, 9, italic=True)
    hdr_row(ws2, 4, [('#', 5), ('golden_id', 34), ('section_type', 20), ('source_type', 15),
                     ('의심 표현', 30), ('판정', 11), ('수정안', 34), ('검토자', 11)])
    for i in range(1, 29):
        put(ws2, 4 + i, [i, '', '', '', '', '', '', ''],
            fills={c: IN for c in (2, 3, 4, 5, 6, 7, 8)})
    dv = DataValidation(type='list', formula1='"승인,수정,폐기,보류"', allow_blank=True, showDropDown=False)
    ws2.add_data_validation(dv); dv.add('F5:F32')
    ws2.freeze_panes = ws2.cell(5, 1)

    # 집계
    ws2.cell(35, 1, '집계').font = Font(F, 11, bold=True)
    for i, (lbl, val) in enumerate([('승인', '승인'), ('수정', '수정'), ('폐기', '폐기'), ('보류', '보류')]):
        ws2.cell(36 + i, 1, lbl).font = Font(F, 10)
        c = ws2.cell(36 + i, 2, f'=COUNTIF(F5:F32,"{val}")')
        c.font = Font(F, 10, bold=True); c.fill = AUTO; c.border = BOX
    ws2.cell(40, 1, '미검토').font = Font(F, 10)
    c = ws2.cell(40, 2, '=28-COUNTA(F5:F32)')
    c.font = Font(F, 10, bold=True); c.fill = AUTO; c.border = BOX

    wb.save(path)
    print(f'saved {path}')


if __name__ == '__main__':
    build_assumptions('가정값_레지스트리.xlsx')
    build_golden_review('Golden_페르소나검토.xlsx')
