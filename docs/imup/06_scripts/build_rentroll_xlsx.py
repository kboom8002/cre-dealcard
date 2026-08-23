# -*- coding: utf-8 -*-
"""
CREDEAL 렌트롤 표준 양식 v1.2 생성기

v1.0 → v1.1 변경
  · '공실'(2값) → '임대상태'(임대중/공실/자가사용) — 자가사용 물건 대응
  · '계약그룹' 컬럼 신설 — 한 임차인이 여러 층을 묶어 쓰는 통합계약 대응
  · '임대면적(평)' 자동 환산 컬럼 신설
  · F13 만료임박(30일) 검증 추가

산출물
  CREDEAL_렌트롤_표준양식_v1.2.xlsx
  CREDEAL_렌트롤_양평동_실측.xlsx
  CREDEAL_렌트롤_양평동_갱신본_가상.xlsx
  CREDEAL_렌트롤_당산동_실측.xlsx      (+ 임대료현실화 시트)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date

F = 'Arial'
C_HDR  = PatternFill('solid', fgColor='1F3864')
C_IN   = PatternFill('solid', fgColor='FFFF00')
C_AUTO = PatternFill('solid', fgColor='E8E8E8')
C_SEC  = PatternFill('solid', fgColor='D9E2F3')
C_WARN = PatternFill('solid', fgColor='FFE0E0')
THIN = Side(style='thin', color='B0B0B0')
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WON, DAT = '#,##0', 'yyyy-mm-dd'
PYEONG = 3.305785

COLS = [
    ('unitLabel',      '호실/층',            10, 'in',   None, 'R1'),
    ('contractGroup',  '계약그룹',           10, 'in',   None, 'R2'),
    ('leaseArea',      '임대면적(㎡)',        12, 'in',   '#,##0.00', 'R2'),
    ('tenantBusiness', '업종/상호 (원문)',    18, 'in',   None, 'R1'),
    ('legalBasis',     '적용법령',            10, 'in',   None, 'R2'),
    ('deposit',        '보증금(원)',          14, 'in',   WON, 'R1'),
    ('monthlyRent',    '월세(원,VAT별도)',    15, 'in',   WON, 'R1'),
    ('mgmtFee',        '관리비(원,VAT별도)',  16, 'in',   WON, 'R2'),
    ('firstContract',  '최초 계약일',         13, 'in',   DAT, 'R3'),
    ('curStart',       '현 계약 시작일',      13, 'in',   DAT, 'R2'),
    ('curExpiry',      '현 계약 만료일',      13, 'in',   DAT, 'R1'),
    ('renewalUsed',    '갱신요구권 행사',     14, 'in',   None, 'R3'),
    ('opposingPower',  '대항력 요건',         13, 'in',   None, 'R3'),
    ('leaseState',     '임대상태',            11, 'in',   None, 'R1'),
    ('note',           '비고',                22, 'in',   None, '—'),
    ('areaPy',         '임대면적(평)',        12, 'auto', '#,##0.0', ''),
    ('convDeposit',    '환산보증금(자동)',    15, 'auto', WON, ''),
    ('sangimApply',    '상임법 전면적용',     14, 'auto', None, ''),
    ('renewRemain',    '갱신권 잔여(자동)',   15, 'auto', None, ''),
    ('expiryState',    '계약 상태(자동)',     15, 'auto', None, ''),
    ('monthlyTotal',   '월 총수입(자동)',     14, 'auto', WON, ''),
]
TIER_FILL = {'R1': 'FFF2CC', 'R2': 'DEEAF6', 'R3': 'FCE4EC', '—': 'F2F2F2', '': 'FFFFFF'}
IDX = {c[0]: i + 1 for i, c in enumerate(COLS)}
def col(k): return get_column_letter(IDX[k])

HDR_ROW = 12
DAT_ROW = HDR_ROW + 1


def sheet_guide(wb):
    ws = wb.create_sheet('기입요령', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 80

    def sec(r, t):
        for c in (2, 3):
            x = ws.cell(r, c, t if c == 2 else '')
            x.font = Font(F, 11, bold=True, color='FFFFFF'); x.fill = C_HDR
    def kv(r, k, v, warn=False):
        a = ws.cell(r, 2, k); a.font = Font(F, 10, bold=True); a.border = BOX
        a.alignment = Alignment(vertical='center')
        b = ws.cell(r, 3, v); b.font = Font(F, 10); b.border = BOX
        b.alignment = Alignment(wrap_text=True, vertical='center')
        if warn: a.fill = C_WARN; b.fill = C_WARN
        ws.row_dimensions[r].height = 30

    ws.cell(2, 2, 'CREDEAL 렌트롤 표준 양식 v1.2 — 기입요령').font = Font(F, 14, bold=True, color='1F3864')
    ws.merge_cells('B2:C2')
    ws.cell(4, 2, '모바일 IM 업로드용 표준 양식입니다. 「렌트롤」 시트의 노란색 셀만 입력하세요.').font = Font(F, 10, italic=True)
    ws.merge_cells('B4:C4')

    r = 6; sec(r, '색상 규칙'); r += 1
    for k, v in [('노란색', '입력 셀 — 여기만 채웁니다'),
                 ('회색', '자동계산 — 수식이 들어 있으니 건드리지 마세요'),
                 ('분홍색', '경고 — 비면 판정이 보류되는 항목')]:
        kv(r, k, v); r += 1

    r += 1; sec(r, '반드시 채워야 하는 항목'); r += 1
    for k, v in [
        ('호실/층', '원본 표기 그대로. 「9F(1)」처럼 한 층에 둘이면 행을 나눠 적습니다.'),
        ('업종/상호 (원문)', '⚠ 계약서·렌트롤에 적힌 문구를 그대로 옮깁니다. 「사무실」이면 「사무실」입니다. '
                            '업종을 추측해 바꾸지 마세요.'),
        ('적용법령', '상가 / 주택 / 미확인. 주거로 쓰는 오피스텔은 「주택」입니다. '
                    '이 값에 따라 갱신요구권 계산이 완전히 달라집니다.'),
        ('임대상태', '임대중 / 공실 / 자가사용. 소유자가 직접 쓰는 층은 「자가사용」입니다 — '
                    '공실이 아니지만 임대수입도 없습니다.'),
        ('보증금·월세·관리비', '원 단위 숫자만. VAT 별도. 「5,000만」처럼 적지 마세요.'),
        ('현 계약 만료일', '명도 시점 산출의 기준입니다.'),
    ]: kv(r, k, v); r += 1

    r += 1; sec(r, '비면 판정이 보류되는 항목'); r += 1
    for k, v in [
        ('최초 계약일', '이 임차인이 처음 입주한 날. 현 계약 시작일과 다릅니다. '
                       '★ 상가 갱신요구권 10년은 이 날짜로 기산합니다. 비면 「확인 필요」로 출력됩니다.'),
        ('갱신요구권 행사', '있음 / 없음 / 모름. ★ 주택은 1회 한정이라 이 값 없이는 계산할 수 없습니다. '
                          '묵시적 갱신은 「없음」입니다 — 갱신요구권을 쓴 것이 아닙니다.'),
        ('대항력 요건', '사업자등록 / 주민등록 / 미확인. 근거 없이 「없음」으로 적으면 발행이 막힙니다.'),
    ]: kv(r, k, v, warn=True); r += 1

    r += 1; sec(r, '계약그룹 — 한 임차인이 여러 층을 쓸 때'); r += 1
    for k, v in [
        ('쓰는 법', '같은 계약이면 같은 이름을 적습니다. 예: 1F와 2F가 한 계약이면 둘 다 「A」.'),
        ('금액 기입', '대표 행에만 보증금·월세를 적고 나머지는 비웁니다. 면적은 각 행에 적습니다.'),
        ('환산보증금', '금액이 있는 대표 행에서만 계산됩니다. 층별로 쪼개지 않습니다.'),
    ]: kv(r, k, v); r += 1

    r += 1; sec(r, '자주 나는 실수'); r += 1
    for k, v in [
        ('업종 추측', '「사무실」을 「IT스타트업」으로 바꿔 적는 경우. 실제로 있었던 오류입니다.'),
        ('경과년수로 갱신권 역산', '주택은 최초계약일로 계산할 수 없습니다. 행사 이력이 있어야 합니다.'),
        ('계 행을 손으로 입력', '합계는 자동입니다. 손으로 적으면 각 행 합과 어긋납니다. 실제로 있었던 오류입니다.'),
        ('자가사용을 공실로', '자가사용은 임대 전환 여력이지 공실 손실이 아닙니다. 구분해 적으세요.'),
        ('환산보증금 직접 계산', '자동입니다. 주택에는 이 개념이 없습니다.'),
    ]: kv(r, k, v); r += 1

    r += 1; sec(r, '자동 계산'); r += 1
    for k, v in [
        ('임대면적(평)', '㎡ ÷ 3.305785'),
        ('환산보증금', '보증금 + 월세×100. 상가·임대중인 대표 행만.'),
        ('상임법 전면적용', '환산보증금이 지역 기준(서울 9억) 이하이면 ○. '
                          '초과하면 5% 인상상한·우선변제권이 적용되지 않습니다.'),
        ('갱신권 잔여', '상가 = 10년 − 최초계약일 경과. 주택 = 행사 이력에 따라 판정.'),
        ('계약 상태', '평가 기준일과 만료일 비교. 30일 이내면 「임박」.'),
        ('월 총수입', '월세 + 관리비. 공실·자가사용은 0.'),
    ]: kv(r, k, v); r += 1
    return ws


def sheet_rentroll(wb, meta, rows, title_txt, name='렌트롤'):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    for i, c in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = c[2]

    ws.cell(1, 1, title_txt).font = Font(F, 15, bold=True, color='1F3864')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)

    for i, (label, val) in enumerate([
        ('물건명', meta.get('name', '')), ('소재지', meta.get('addr', '')),
        ('렌트롤 기준일', meta.get('asOf')), ('평가 기준일', meta.get('evalAt')),
        ('환산보증금 지역기준(원)', meta.get('threshold', 900000000)),
        ('작성자', meta.get('author', '')), ('작성일', meta.get('writtenAt')),
    ]):
        r = 3 + i
        a = ws.cell(r, 1, label); a.font = Font(F, 10, bold=True); a.fill = C_SEC; a.border = BOX
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        b = ws.cell(r, 3, val); b.font = Font(F, 10); b.fill = C_IN; b.border = BOX
        if label.endswith('일'): b.number_format = DAT
        if '지역기준' in label: b.number_format = WON
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws.cell(3, 6, '⚠ 렌트롤 기준일이 비면 신선도 판정을 할 수 없습니다').font = Font(F, 9, italic=True, color='C00000')
    ws.cell(4, 6, '※ 환산보증금: 서울 9억 · 과밀억제권역 6.9억 · 광역시 5.4억 · 그 밖 3.7억').font = Font(F, 9, italic=True, color='555555')

    # 해상도 등급 띠 (헤더 바로 위)
    ws.cell(HDR_ROW - 1, 1, '해상도').font = Font(F, 8, bold=True, color='555555')
    for i, (k, h, w, kind, fmt, tg) in enumerate(COLS, 1):
        if i == 1: continue
        t = ws.cell(HDR_ROW - 1, i, tg)
        t.font = Font(F, 8, bold=True, color='555555')
        t.fill = PatternFill('solid', fgColor=TIER_FILL.get(tg, 'FFFFFF'))
        t.alignment = Alignment(horizontal='center')
        t.border = BOX
    ws.row_dimensions[HDR_ROW - 1].height = 14

    for i, (k, h, w, kind, fmt, tg) in enumerate(COLS, 1):
        c = ws.cell(HDR_ROW, i, h)
        c.font = Font(F, 9, bold=True, color='FFFFFF'); c.fill = C_HDR
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BOX
    ws.row_dimensions[HDR_ROW].height = 32
    ws.freeze_panes = ws.cell(DAT_ROW, 1)

    ev, th = '$C$6', '$C$7'
    for j, row in enumerate(rows):
        rr = DAT_ROW + j
        for i, (k, h, w, kind, fmt, tg) in enumerate(COLS, 1):
            c = ws.cell(rr, i); c.border = BOX; c.font = Font(F, 9)
            c.alignment = Alignment(horizontal='center', vertical='center')
            if fmt: c.number_format = fmt
            c.fill = C_IN if kind == 'in' else C_AUTO
            if kind == 'in' and k in row: c.value = row[k]

        D, M, MG = col('deposit'), col('monthlyRent'), col('mgmtFee')
        LB, ST = col('legalBasis'), col('leaseState')
        FC, EX, RU, AR = col('firstContract'), col('curExpiry'), col('renewalUsed'), col('leaseArea')

        ws[f'{col("areaPy")}{rr}'] = f'=IF({AR}{rr}="","",{AR}{rr}/{PYEONG})'
        ws[f'{col("convDeposit")}{rr}'] = (
            f'=IF(OR({ST}{rr}<>"임대중",{LB}{rr}<>"상가",{D}{rr}=""),"",{D}{rr}+{M}{rr}*100)')
        ws[f'{col("sangimApply")}{rr}'] = (
            f'=IF({col("convDeposit")}{rr}="","",'
            f'IF({col("convDeposit")}{rr}<={th},"○ 전면적용","× 초과 (5%상한 미적용)"))')
        ws[f'{col("renewRemain")}{rr}'] = (
            f'=IF({ST}{rr}<>"임대중","",'
            f'IF({LB}{rr}="상가",IF({FC}{rr}="","확인 필요",'
            f'TEXT(MAX(0,10-({ev}-{FC}{rr})/365.25),"0.0")&"년"),'
            f'IF({LB}{rr}="주택",IF({RU}{rr}="있음","소진(1회)",'
            f'IF({RU}{rr}="없음",IF({EX}{rr}="","확인 필요",TEXT(EDATE({EX}{rr},24),"yyyy-mm")&"까지"),"확인 필요")),'
            f'"확인 필요")))')
        ws[f'{col("expiryState")}{rr}'] = (
            f'=IF({ST}{rr}="공실","공실",IF({ST}{rr}="자가사용","자가사용",'
            f'IF({EX}{rr}="","만료일 없음",'
            f'IF({EX}{rr}<{ev},"만료 +"&TEXT({ev}-{EX}{rr},"0")&"일",'
            f'IF({EX}{rr}-{ev}<=30,"임박 "&TEXT({EX}{rr}-{ev},"0")&"일","유효")))))')
        ws[f'{col("monthlyTotal")}{rr}'] = (
            f'=IF({ST}{rr}<>"임대중",0,IF({M}{rr}="",0,{M}{rr})+IF({MG}{rr}="",0,{MG}{rr}))')
        ws.row_dimensions[rr].height = 18

    tot = DAT_ROW + len(rows)
    c = ws.cell(tot, 1, '합계'); c.font = Font(F, 10, bold=True); c.fill = C_SEC; c.border = BOX
    ws.merge_cells(start_row=tot, start_column=1, end_row=tot, end_column=2)
    for k in ('leaseArea', 'deposit', 'monthlyRent', 'mgmtFee', 'areaPy', 'monthlyTotal'):
        cc = ws.cell(tot, IDX[k], f'=SUM({col(k)}{DAT_ROW}:{col(k)}{tot-1})')
        cc.font = Font(F, 10, bold=True); cc.fill = C_SEC; cc.border = BOX
        cc.number_format = dict(COLS[IDX[k]-1:IDX[k]])[list(dict(COLS[IDX[k]-1:IDX[k]]).keys())[0]] if False else \
            ('#,##0.00' if k == 'leaseArea' else '#,##0.0' if k == 'areaPy' else WON)
        cc.alignment = Alignment(horizontal='center')
    for i in range(3, len(COLS) + 1):
        if ws.cell(tot, i).value is None:
            ws.cell(tot, i).fill = C_SEC; ws.cell(tot, i).border = BOX

    for k, f_ in [('legalBasis', '"상가,주택,미확인"'),
                  ('renewalUsed', '"있음,없음,모름"'),
                  ('opposingPower', '"사업자등록,주민등록,미확인"'),
                  ('leaseState', '"임대중,공실,자가사용"')]:
        dv = DataValidation(type='list', formula1=f_, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv); dv.add(f'{col(k)}{DAT_ROW}:{col(k)}{DAT_ROW+199}')
    return ws, tot


def sheet_check(wb, tot_row, n_rows, sheet='렌트롤'):
    ws = wb.create_sheet('자동검증')
    ws.sheet_view.showGridLines = False
    for c, w in zip('ABCDE', (3, 6, 32, 22, 48)):
        ws.column_dimensions[c].width = w
    ws.cell(2, 2, '자동 검증 — 발행 전 확인').font = Font(F, 14, bold=True, color='1F3864')
    ws.merge_cells('B2:E2')
    for i, h in enumerate(['#', '검증 항목', '결과', '판정 기준'], 2):
        c = ws.cell(4, i, h); c.font = Font(F, 10, bold=True, color='FFFFFF')
        c.fill = C_HDR; c.border = BOX; c.alignment = Alignment(horizontal='center')

    R, d0, d1 = sheet, DAT_ROW, tot_row - 1
    ES, US = col('expiryState'), col('unitLabel')
    checks = [
        ('G19', '렌트롤 월세 합계 (정본)',
         f"=TEXT('{R}'!{col('monthlyRent')}{tot_row},\"#,##0\")&\"원\"",
         '표지 요약과 다르면 발행 차단. 이 표가 정본입니다.'),
        ('C19', '임대면적 합계',
         f"=TEXT('{R}'!{col('leaseArea')}{tot_row},\"#,##0.00\")&\"㎡ / \"&TEXT('{R}'!{col('areaPy')}{tot_row},\"#,##0.0\")&\"평\"",
         '건축물대장 연면적과 대조. 불일치 시 원인 규명 전까지 발행 보류.'),
        ('F11', '만료 계약 수',
         f"=COUNTIF('{R}'!{ES}{d0}:{ES}{d1},\"만료*\")&\" / \"&COUNTIF('{R}'!{col('leaseState')}{d0}:{col('leaseState')}{d1},\"임대중\")&\"건\"",
         '1건 이상이면 호실별 경고.'),
        ('F12', '만료 비율 50% 초과',
         f"=IF(COUNTIF('{R}'!{ES}{d0}:{ES}{d1},\"만료*\")>"
         f"COUNTIF('{R}'!{col('leaseState')}{d0}:{col('leaseState')}{d1},\"임대중\")/2,\"발행 차단\",\"통과\")",
         '절반 초과 시 렌트롤 갱신 전까지 발행 불가.'),
        ('F13', '30일 내 만료 임박',
         f"=COUNTIF('{R}'!{ES}{d0}:{ES}{d1},\"임박*\")&\"건\"",
         '재계약 조건이 딜 구조를 바꿉니다. 협의 상태 확인 필요.'),
        ('G18', '갱신권 판정 보류',
         f"=COUNTIF('{R}'!{col('renewRemain')}{d0}:{col('renewRemain')}{d1},\"확인 필요\")&\"건\"",
         '최초계약일(상가)·행사이력(주택) 누락. 숫자 대신 "확인 필요" 출력.'),
        ('G13', '대항력 미확인',
         f"=COUNTIF('{R}'!{col('opposingPower')}{d0}:{col('opposingPower')}{d1},\"미확인\")&\"건\"",
         '근거 없이 "없음" 표기 시 발행 차단.'),
        ('G17', '업종 미기재',
         f"=COUNTBLANK('{R}'!{col('tenantBusiness')}{d0}:{col('tenantBusiness')}{d1})&\"건\"",
         '비면 "미상"으로 출력. 추론 금지.'),
        ('T-C', '환산보증금 초과 호실',
         f"=COUNTIF('{R}'!{col('sangimApply')}{d0}:{col('sangimApply')}{d1},\"×*\")&\"건\"",
         '초과 시 5% 인상상한·우선변제권 미적용. 갱신요구권은 적용됨.'),
        ('F01', '렌트롤 기준일',
         f"=IF('{R}'!$C$5=\"\",\"미기재\",TEXT('{R}'!$C$5,\"yyyy-mm-dd\"))",
         '없으면 신선도 판정 불가.'),
        ('C21', '공실 / 자가사용',
         f"=COUNTIF('{R}'!{col('leaseState')}{d0}:{col('leaseState')}{d1},\"공실\")&\"건 / \"&"
         f"COUNTIF('{R}'!{col('leaseState')}{d0}:{col('leaseState')}{d1},\"자가사용\")&\"건\"",
         '자가사용은 임대 전환 여력. 공실 손실과 구분합니다.'),
    ]
    r = 5
    for code, name, formula, basis in checks:
        ws.cell(r, 2, code).font = Font(F, 9, bold=True)
        ws.cell(r, 3, name).font = Font(F, 10)
        cc = ws.cell(r, 4, formula); cc.font = Font(F, 10, bold=True); cc.fill = C_AUTO
        ws.cell(r, 5, basis).font = Font(F, 9, color='555555')
        for i in range(2, 6):
            ws.cell(r, i).border = BOX
            ws.cell(r, i).alignment = Alignment(vertical='center', wrap_text=(i == 5))
        for i in (2, 4):
            ws.cell(r, i).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 26
        r += 1

    # ── 해상도 판정 블록 ──
    r += 1
    ws.cell(r, 2, '해상도 판정 — 이 렌트롤로 무엇을 만들 수 있나').font = Font(F, 13, bold=True, color='1F3864')
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    for i, h in enumerate(['등급', '요건', '충족', '열리는 기능'], 2):
        c = ws.cell(r, i, h); c.font = Font(F, 10, bold=True, color='FFFFFF')
        c.fill = C_HDR; c.border = BOX; c.alignment = Alignment(horizontal='center')
    r += 1

    LIVE = f"COUNTIF('{R}'!{col('leaseState')}{d0}:{col('leaseState')}{d1},\"임대중\")"
    def filled(k):   # 임대중 행 중 값이 있는 개수
        return (f"COUNTIFS('{R}'!{col('leaseState')}{d0}:{col('leaseState')}{d1},\"임대중\","
                f"'{R}'!{col(k)}{d0}:{col(k)}{d1},\"<>\")")
    def allrows(k):  # 전체 데이터 행 중 값이 있는 개수
        return f"COUNTA('{R}'!{col(k)}{d0}:{col(k)}{d1})"
    NROW = f"COUNTA('{R}'!{col('unitLabel')}{d0}:{col('unitLabel')}{d1})"

    r1 = f"AND({filled('tenantBusiness')}={LIVE},{filled('curExpiry')}={LIVE},{allrows('leaseState')}={NROW})"
    r2 = f"AND({r1},{allrows('leaseArea')}={NROW},{allrows('legalBasis')}={NROW},{filled('mgmtFee')}={LIVE})"
    r3 = f"AND({r2},{filled('firstContract')}={LIVE},"
    r3 += f"COUNTIFS('{R}'!{col('leaseState')}{d0}:{col('leaseState')}{d1},\"임대중\","
    r3 += f"'{R}'!{col('opposingPower')}{d0}:{col('opposingPower')}{d1},\"미확인\")=0)"

    tiers = [
        ('R1 최소형', '호실·업종·금액·만료일·임대상태', r1,
         '렌트롤 표 · 만료 타임라인 · 총액 수익률'),
        ('R2 필요형', 'R1 + 면적·적용법령·관리비·시작일', r2,
         '평당 단가 · 층별 단가 · 환산보증금 · 상임법 판정'),
        ('R3 표준형', 'R2 + 최초계약일·대항력', r3,
         '갱신요구권 · 명도 계획 · 임대료 정상화 시뮬'),
    ]
    for nm, req, cond, unlock in tiers:
        ws.cell(r, 2, nm).font = Font(F, 10, bold=True)
        ws.cell(r, 3, req).font = Font(F, 9)
        c = ws.cell(r, 4, f'=IF({cond},"충족","미달")')
        c.font = Font(F, 10, bold=True); c.fill = C_AUTO
        ws.cell(r, 5, unlock).font = Font(F, 9, color='555555')
        for i in range(2, 6):
            ws.cell(r, i).border = BOX
            ws.cell(r, i).alignment = Alignment(vertical='center', wrap_text=(i in (3, 5)))
        ws.cell(r, 4).alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    ws.cell(r, 2, '현재 해상도').font = Font(F, 11, bold=True)
    cc = ws.cell(r, 3, f'=IF({r3},"R3 표준형",IF({r2},"R2 필요형",IF({r1},"R1 최소형","R0 미달")))')
    cc.font = Font(F, 14, bold=True, color='1F3864'); cc.fill = C_AUTO; cc.border = BOX
    cc.alignment = Alignment(horizontal='center')
    r += 1
    ws.cell(r, 2, '다음 한 칸').font = Font(F, 11, bold=True)
    nxt = (f'=IF({r3},"완료 — 모든 기능 사용 가능",'
           f'IF({r2},"「최초 계약일」과 「대항력 요건」을 채우면 갱신요구권·명도 계획이 열립니다",'
           f'IF({r1},"「적용법령」을 채우면 환산보증금·상임법 판정이 열립니다. 「임대면적」은 평당 단가를 엽니다",'
           f'"「업종」과 「현 계약 만료일」부터 채우세요. 렌트롤 표를 그릴 수 없습니다")))')
    c2 = ws.cell(r, 3, nxt); c2.font = Font(F, 10, color='C00000')
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    c2.alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[r].height = 32
    return ws


def build(path, meta, rows, title_txt, n_blank=0, plan=None, plan_meta=None, plan_title=''):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    sheet_guide(wb)
    allrows = list(rows) + [{} for _ in range(n_blank)]
    _, tot = sheet_rentroll(wb, meta, allrows, title_txt)
    if plan:
        sheet_rentroll(wb, plan_meta or meta, plan, plan_title, name='임대료현실화')
    sheet_check(wb, tot, len(allrows))
    wb.save(path)
    print(f'saved {path}  ({len(allrows)}행, 합계 {tot}행)')


# ══════════════════════════════════════════════════════════════
if __name__ == '__main__':
    EV = date(2026, 8, 23)
    SEOUL = 900000000

    # ─── ① 빈 템플릿 ───
    build('CREDEAL_렌트롤_표준양식_v1.2.xlsx',
          {'name': '', 'addr': '', 'asOf': None, 'evalAt': EV,
           'threshold': SEOUL, 'author': '', 'writtenAt': None},
          [{'unitLabel': '3F', 'leaseArea': 132.5, 'tenantBusiness': '사무실',
            'legalBasis': '상가', 'deposit': 50000000, 'monthlyRent': 4500000,
            'mgmtFee': 550000, 'firstContract': date(2019, 4, 1),
            'curStart': date(2025, 4, 1), 'curExpiry': date(2027, 3, 31),
            'renewalUsed': '모름', 'opposingPower': '사업자등록', 'leaseState': '임대중',
            'note': '← 예시 행입니다. 지우고 쓰세요'}],
          'CREDEAL 렌트롤 표준 양식 v1.2', n_blank=24)

    # ─── ② 양평동 실측 ───
    def Y(u, biz, dep, rent, mgmt, s, e, area=None, note=''):
        return {'unitLabel': u, 'leaseArea': area, 'tenantBusiness': biz,
                'legalBasis': '상가', 'deposit': dep, 'monthlyRent': rent, 'mgmtFee': mgmt,
                'firstContract': None, 'curStart': s, 'curExpiry': e,
                'renewalUsed': '모름', 'opposingPower': '미확인',
                'leaseState': '임대중', 'note': note}
    yp = [
        Y('10F','운동시설',    43000000,4590000,510000, date(2022,10,1), date(2024,9,30)),
        Y('9F(2)','사무실',    30000000,3000000,430000, date(2023,12,1), date(2025,11,30)),
        Y('9F(1)','스튜디오렌탈',20000000,1990000,220000, date(2023,11,1), date(2025,10,31)),
        Y('8F','사무실',       50000000,5600000,800000, date(2023,9,8),  date(2025,9,7)),
        Y('7F','사무실',       50000000,5880000,620000, date(2023,10,4), date(2025,10,3)),
        Y('6F','사무실',       40000000,4830000,690000, date(2024,3,1),  date(2026,2,28)),
        Y('5F','사무실',       57000000,5280000,660000, date(2022,5,30), date(2024,5,29)),
        Y('4F','사무실',       50000000,4400000,550000, date(2022,6,30), date(2024,6,29)),
        Y('3F','치과',        70000000,3100000,530000, date(2022,11,1), date(2024,10,31)),
        Y('2F','미용실',       50000000,5400000,600000, date(2024,2,22), date(2026,2,21)),
        Y('1F','부동산',       35000000,2500000,150000, date(2023,11,11),date(2025,11,11)),
        {'unitLabel':'B1','leaseArea':422.25,'tenantBusiness':'','legalBasis':'미확인',
         'leaseState':'공실','renewalUsed':'모름','opposingPower':'미확인',
         'note':'공실 · 리스업 대상'},
    ]
    build('CREDEAL_렌트롤_양평동_실측.xlsx',
          {'name':'더레드빌딩','addr':'서울특별시 영등포구 양평동4가 117, 134, 125-2',
           'asOf':None,'evalAt':EV,'threshold':SEOUL,
           'author':'(원본 IM 내장 이미지 판독)','writtenAt':EV},
          yp, 'CREDEAL 렌트롤 — 양평동4가 117 (실측)')

    # ─── ③ 양평동 갱신본 (가상) ───
    def V(u, biz, dep, rent, mgmt, first, s, e):
        return {'unitLabel':u,'tenantBusiness':biz,'legalBasis':'상가','deposit':dep,
                'monthlyRent':rent,'mgmtFee':mgmt,'firstContract':first,'curStart':s,
                'curExpiry':e,'renewalUsed':'모름','opposingPower':'사업자등록',
                'leaseState':'임대중','note':'가상 갱신본'}
    yp2 = [
        V('10F','운동시설',    43000000,4590000,510000, date(2018,10,1), date(2024,10,1), date(2026,9,30)),
        V('9F(2)','사무실',    30000000,3000000,430000, date(2021,12,1), date(2025,12,1), date(2027,11,30)),
        V('9F(1)','스튜디오렌탈',20000000,1990000,220000, date(2019,11,1), date(2025,11,1), date(2027,10,31)),
        V('8F','사무실',       50000000,5600000,800000, date(2019,9,8),  date(2025,9,8),  date(2027,9,7)),
        V('7F','사무실',       50000000,5880000,620000, date(2021,10,4), date(2025,10,4), date(2027,10,3)),
        V('6F','사무실',       40000000,4830000,690000, date(2022,3,1),  date(2026,3,1),  date(2028,2,28)),
        V('5F','사무실',       57000000,5280000,660000, date(2018,10,15),date(2026,5,30), date(2028,5,29)),
        V('4F','사무실',       50000000,4400000,550000, date(2018,11,1), date(2026,6,30), date(2028,6,29)),
        V('3F','치과',        70000000,3100000,530000, date(2018,11,1), date(2024,11,1), date(2026,10,31)),
        V('2F','미용실',       50000000,5400000,600000, date(2020,2,22), date(2026,2,22), date(2028,2,21)),
        V('1F','부동산',       35000000,2500000,150000, date(2018,11,11),date(2025,11,11),date(2027,11,10)),
        {'unitLabel':'B1','leaseArea':422.25,'tenantBusiness':'','legalBasis':'미확인',
         'leaseState':'공실','renewalUsed':'모름','opposingPower':'미확인','note':'공실'},
    ]
    build('CREDEAL_렌트롤_양평동_갱신본_가상.xlsx',
          {'name':'더레드빌딩 (가상 갱신본)',
           'addr':'서울특별시 영등포구 양평동4가 117, 134, 125-2',
           'asOf':date(2026,8,1),'evalAt':EV,'threshold':SEOUL,
           'author':'⚠ 테스트 전용 가상 데이터 — 실제 임대차 아님','writtenAt':EV},
          yp2, 'CREDEAL 렌트롤 — 양평동 (가상 갱신본 · 테스트 전용)')

    # ─── ④ 당산동 실측 (+ 임대료현실화) ───
    K = 1000
    def D_(u, grp, area, biz, dep, rent, e, state='임대중', note=''):
        return {'unitLabel':u,'contractGroup':grp,'leaseArea':area,'tenantBusiness':biz,
                'legalBasis':'상가' if state=='임대중' else '미확인',
                'deposit':dep*K if dep else None,'monthlyRent':rent*K if rent else None,
                'mgmtFee':None,'firstContract':None,'curStart':None,'curExpiry':e,
                'renewalUsed':'모름','opposingPower':'미확인','leaseState':state,'note':note}
    E831 = date(2026,8,31)
    ds = [
        D_('B1','',317.22,'데이르 카페',None,None,None,'자가사용','소유자 직접 운영 · 임대 전환 대상'),
        D_('1F','A',78.39,'고은약국',60000,1830,E831,note='비고 "임대 11년 경과"'),
        D_('1F','B',105.60,'로뎀나무내과',140000,8830,E831,note='1F+2F 통합계약 · "임대 11년 경과"'),
        D_('2F','B',252.09,'로뎀나무내과',None,None,E831,note='B그룹 (금액은 1F 행)'),
        D_('3F','C',252.09,'헬쓰장',50000,4550,date(2026,4,17)),
        D_('4F','D',169.06,'국제와인',30000,2600,date(2025,4,30),note='IM 작성(2025.05) 시점에 이미 만료'),
        D_('4F','',83.03,'(자가)',None,None,None,'자가사용','임대 전환 대상'),
        D_('5F','E',183.67,'로뎀나무내과',10000,1650,E831,note='별도 계약 · "임대 11년 경과"'),
    ]
    plan = [
        D_('B1','P1',317.22,'데이르 카페',50000,4580,None,'임대중','자가 → 임대 전환 · 기준층의 80%'),
        D_('1F','P2',78.39,'고은약국',60000,3450,None,'임대중','기준층의 2.5배'),
        D_('1F','P3',105.60,'로뎀나무내과',140000,8160,None,'임대중','1F+2F 통합 · 기준층의 2배 · 현행 8,830 → 감액'),
        D_('2F','P3',252.09,'로뎀나무내과',None,None,None,'임대중','P3그룹 · 기준층 동일'),
        D_('3F','P4',252.09,'헬쓰장',50000,4550,None,'임대중','기준단가 62.4천원/평 · 현행 유지'),
        D_('4F','P5',169.06,'국제와인',30000,3060,None,'임대중','기준층 동일 · 갱신요구권 잔여 확인 필요'),
        D_('4F','P6',83.03,'(자가)',30000,1440,None,'임대중','자가 → 임대 전환'),
        D_('5F','P7',183.67,'로뎀나무내과',10000,3430,None,'임대중','기준층 동일'),
    ]
    build('CREDEAL_렌트롤_당산동_실측.xlsx',
          {'name':'당산동5가 11-47 근생빌딩',
           'addr':'서울특별시 영등포구 당산동5가 11-47',
           'asOf':date(2025,5,1),'evalAt':EV,'threshold':SEOUL,
           'author':'(원본 IM p6 판독)','writtenAt':EV},
          ds, 'CREDEAL 렌트롤 — 당산동5가 11-47 (실측)',
          plan=plan,
          plan_meta={'name':'당산동5가 11-47 — 임대료 현실화 계획',
                     'addr':'서울특별시 영등포구 당산동5가 11-47',
                     'asOf':date(2025,5,1),'evalAt':EV,'threshold':SEOUL,
                     'author':'(원본 IM p7 판독) ◇가정 — 실현 미확정','writtenAt':EV},
          plan_title='임대료 현실화 계획 (원본 p7) — ◇가정')
